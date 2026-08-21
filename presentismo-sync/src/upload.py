"""Sube el Excel de "Detalle de marcas" a Supabase.

Misma lógica que la sección "Presentismo WM — Marcaciones (DATA)" de
admin-panel.html (y que el bot Node eliminado) — portada a Python. Usa la
service_role key así que no pasa por RLS.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from supabase import Client, create_client

BATCH = 1000


def _clean_text(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _parse_fecha_hora(v: Any) -> dt.datetime | None:
    """El Excel no trae huso horario — la hora es la hora de Chile tal
    cual. openpyxl ya devuelve un datetime naive para celdas con formato
    de fecha; se trata ese naive tal cual como si fuera UTC (mismo
    criterio que el bot anterior en JS: Date.UTC con los mismos números,
    no la hora local de quien corre el script)."""
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, str) and v.strip():
        s = v.strip().replace(" ", "T")
        try:
            return dt.datetime.fromisoformat(s)
        except ValueError:
            return None
    return None


def _read_detalle_rows(file_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(file_path, data_only=True)
    if "Detalle" not in wb.sheetnames:
        raise ValueError(f'El archivo no tiene una hoja "Detalle" (hojas: {", ".join(wb.sheetnames)})')
    ws = wb["Detalle"]

    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]

    rows = []
    for raw_row in rows_iter:
        row = dict(zip(header, raw_row))
        if all(v is None for v in row.values()):
            continue
        rows.append(row)
    return rows


def upload_presentismo_file(*, file_path: Path, supabase_url: str, supabase_service_key: str) -> dict[str, int]:
    raw = _read_detalle_rows(file_path)

    rows = []
    for r in raw:
        entrada = _parse_fecha_hora(r.get("ENTRADA"))
        salida = _parse_fecha_hora(r.get("SALIDA"))
        formato = _clean_text(r.get("FORMATO"))
        # Formato SBA se descarta a pedido explícito — no corresponde a
        # las salas que trackea Presentismo WM.
        if entrada is None or r.get("LOCAL") is None or not r.get("RUT PERSONA"):
            continue
        if formato is not None and formato.upper() == "SBA":
            continue
        rows.append({**r, "_entrada": entrada, "_salida": salida})

    supabase: Client = create_client(supabase_url, supabase_service_key)

    salas_resp = supabase.table("salas").select("id, local_code").not_.is_("local_code", "null").execute()
    sala_by_local = {int(s["local_code"]): s["id"] for s in salas_resp.data}

    upsert_rows = []
    sin_sala = 0
    for r in rows:
        local = int(r["LOCAL"])
        sala_id = sala_by_local.get(local)
        if sala_id is None:
            sin_sala += 1
        upsert_rows.append(
            {
                "sala_id": sala_id,
                "local_code": local,
                "nombre_local": _clean_text(r.get("NOMBRE LOCAL")),
                "formato": _clean_text(r.get("FORMATO")),
                "rut_persona": str(r["RUT PERSONA"]).strip(),
                "nombre_persona": _clean_text(r.get("NOMBRE PERSONA")),
                "cargo": _clean_text(r.get("CARGO")),
                "entrada": r["_entrada"].isoformat() + "Z",
                "salida": (r["_salida"].isoformat() + "Z") if r["_salida"] else None,
            }
        )

    for i in range(0, len(upsert_rows), BATCH):
        batch = upsert_rows[i : i + BATCH]
        supabase.table("presentismo_registros").upsert(batch, on_conflict="rut_persona,local_code,entrada").execute()

    return {"total": len(rows), "cargadas": len(upsert_rows), "sin_sala": sin_sala}
